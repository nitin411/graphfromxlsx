
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
from pathlib import Path
import matplotlib.ticker as tick
import matplotlib

def y_fmt(y, pos):
    decades = [1e9, 1e6, 1e3, 1e0, 1e-3, 1e-6, 1e-9 ]
    suffix  = ["G", "M", "k", "" , "m" , "u", "n"  ]
    if y == 0:
        return str(0)
    for i, d in enumerate(decades):
        if np.abs(y) >=d:
            val = y/float(d)
            signf = len(str(val).split(".")[1])
            if signf == 0:
                return '{val:d} {suffix}'.format(val=int(val), suffix=suffix[i])
            else:
                if signf == 1:
                    if str(val).split(".")[1] == "0":
                        return '{val:d} {suffix}'.format(val=int(round(val)), suffix=suffix[i])
                tx = "{"+"val:.{signf}f".format(signf = signf) +"} {suffix}"
                return tx.format(val=val, suffix=suffix[i])
    return y


def get_xlsx_file(path, company):
    company = company.lower().strip()
    paths = Path(path).glob('*.xlsx')
    for path in paths:
        if company in str(path).lower():
            return path


def load_xlsx(path, company_name, heading, subheading):
    date_flag = False
    data = {}
    for company in company_name:
        complete_flag = False
        xlsx_file = get_xlsx_file(path, company)
        wb = openpyxl.load_workbook(xlsx_file)
        sheet=wb['Data Sheet']
        col_names = []
        for column in sheet.iter_cols(1,sheet.max_column):
            col_names.append(column[0].value)
        max_row=sheet.max_row

        max_column=sheet.max_column
        for i in range(1,max_row+1):
            cell = sheet.cell(row=i, column =1).value
            if cell and cell.lower().strip() == heading.lower().strip():
                i = i + 1
                if date_flag == False:
                    date = []
                    for p in range(2, max_column+1):
                        date.append(sheet.cell(row=i, column =p).value.strftime("%b-%Y"))
                    data['Date'] = date
                    date_flag = True
                for j in range(i, max_row+1):
                    if str(sheet.cell(row=j, column=1).value).lower().strip() == subheading.lower().strip():
                        metrics = []
                        for k in range(2, max_column+1):
                            metrics.append(sheet.cell(row=j, column =k).value)
                        data[company] = metrics
                        break
            if complete_flag:
                break
    return data


def addlabels(x,y):
    for i in range(len(x)):
        plt.text(i, y[i], y[i], ha = 'center')

def plot(data):
    matplotlib.style.use('fivethirtyeight')
    data = pd.DataFrame(data)
    data.set_index('Date',inplace=True)
    fig, ax = plt.subplots(figsize=(14,10), facecolor='#A8FECE')
    fig.text(0.50, 0.50, '©iTimes',
             fontsize=80, color='gray',
             ha='center', va='center', alpha=0.4)

    data.plot.bar(rot=0, ax = ax, width = 0.5)
    ax.set_title('iTimes', fontsize=26)
    ax.set_ylabel('Sales', fontsize=26)
    ax.set_xlabel('Date', fontsize=26)
    ax.yaxis.set_major_formatter(tick.FuncFormatter(y_fmt))
    ax.patch.set_facecolor('white')
    for label in (ax.get_xticklabels() + ax.get_yticklabels()):
        label.set_fontsize(16)
    plt.legend(loc=2, prop={'size': 20})
    fig.autofmt_xdate()
    print(data)
    plt.show()


if __name__ == '__main__':
    company_name = [ 'reliance', 'rain industries']
    path = "<add_the_directory_path_here>"
    path='/Users/nshrivastava/Downloads'
    data = load_xlsx(path, company_name, 'profit & loss', 'sales')
    plot(data)